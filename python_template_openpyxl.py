#! /usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import io
import logging
import datetime
from openpyxl import load_workbook
from dataclasses import dataclass, field

# cSpell:ignore openpyxl datefmt levelname dataclass

# get logger of this library __name__ and attach a null handler
logger = logging.getLogger(__name__)
logger.addHandler(logging.NullHandler())


@dataclass
class Report:
    file   : str
    records: list = field(default_factory=list, init=False)

    def __post_init__(self):
        pass


@dataclass
class Record:
    row_number: int
    some_str  : str
    some_int  : int
    some_float: float
    some_date : datetime.date

    def __post_init__(self):
        pass


def parse(file, worksheet, row_of_label, callback=None):

    # iterate through rows of an Excel spreadsheet

    if callback is not None and not callable(callback):
        raise ValueError(
            f'callback given but it it not callable. It is a {type(callback)}.'
        )

    wb = load_workbook(filename=file, read_only=True, data_only=True)
    ws = wb[worksheet]
    logger.debug(f'Slurp item information. Looking at worksheet "{ws.title}".')

    r = Report(file=file)

    logger.debug(f'Max row of this sheet is {ws.max_row}.')
    for row_number, row in enumerate(ws.iter_rows(min_row=row_of_label+1, max_row=ws.max_row+1, values_only=True), start=row_of_label+1):
        # In openpyxl, row and column are one based, not zero based. Cell "A1" is row 1, col 1.
        # Note that when you iterate through rows, now a row is a tuple, and it is zero based.
        # In this example, row[0] is column A.
        # a note on cell value evaluation:
        # the value of a blank cell is None.
        # the value of zero length string (such as ="") is None
        # if you set data_only True on your load_workbook().
        # the value of zero length string (such as ="") is '=""'
        # if you do not set data_only property True on your load_workbook().
        r.records.append(
            Record(
                row_number = row_number,
                some_str   = row[0],
                some_int   = row[1],
                some_float = row[2],
                some_date  = row[3],
            )
        )

        if callback is not None and callable(callback):
            callback(r.records[-1])

    if callback is None:
        return r


def main():
    pass


if __name__ == "__main__":

    # https://qiita.com/jack-low/items/91bf9b5342965352cbeb
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

    # logger setup:
    # if this library is run as a script, these logger setup is used
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)

    file_name = str(sys.argv[0])[:-3] + ".log"
    handler_file = logging.FileHandler(file_name)
    handler_file.setLevel(logging.DEBUG)
    formatter_file = logging.Formatter(
        "%(asctime)s - %(filename)s: %(lineno)s: %(funcName)s - %(levelname)s: %(message)s"
    )
    handler_file.setFormatter(formatter_file)
    logger.addHandler(handler_file)

    # console logger to show INFO messages
    handler_console = logging.StreamHandler()
    handler_console.setLevel(logging.INFO)
    formatter_console = logging.Formatter("%(name)s: %(levelname)s %(message)s")
    handler_console.setFormatter(formatter_console)
    logger.addHandler(handler_console)

    main()
